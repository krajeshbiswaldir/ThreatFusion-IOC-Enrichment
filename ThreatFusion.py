#!/usr/bin/env python3
"""
ThreatFusion - IOC Enrichment Tool
A professional SOC-grade command-line tool for enriching Indicators of Compromise (IOCs)
using VirusTotal, AbuseIPDB, AlienVault OTX, WHOIS, GeoIP, and Shodan.

Author: ThreatFusion Engine
Version: 1.0.0
Python: 3.12+
"""

import os
import re
import csv
import json
import time
import hashlib
import logging
import socket
import urllib.parse
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from typing import Optional, List, Dict, Any, Tuple

import requests
import pandas as pd
import pycountry
import whois
import shodan
from tqdm import tqdm
from jinja2 import Template
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, ListFlowable, ListItem
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# API CONFIGURATION
# =============================================================================
VT_API_KEY = ""
ABUSE_API_KEY = ""
OTX_API_KEY = ""
SHODAN_API_KEY = ""
GEOIP_API_KEY = ""  # Optional: ip-api.com is used as primary (free, no key)

# =============================================================================
# CONSTANTS
# =============================================================================
INPUT_FILE = "IOC List.txt"
OUTPUT_DIR = "Output"
LOG_FILE = "logs.txt"
MAX_WORKERS = 5
REQUEST_TIMEOUT = 30
RETRY_ATTEMPTS = 3

# =============================================================================
# LOGGING SETUP
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("ThreatFusion")


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class EnrichedIOC:
    """Container for enriched IOC data across all supported types."""
    ioc: str
    ioc_type: str

    # IP-specific fields
    country: str = ""
    city: str = ""
    asn: str = ""
    isp: str = ""
    network: str = ""
    latitude: str = ""
    longitude: str = ""

    # Domain-specific fields
    registrar: str = ""
    creation_date: str = ""
    expiration_date: str = ""

    # URL-specific fields
    final_url: str = ""
    status: str = ""
    category: str = ""

    # Hash-specific fields
    file_type: str = ""
    magic: str = ""
    names: str = ""
    malware_family: str = ""
    tags: str = ""

    # Common enrichment fields
    vt_score: str = ""
    vt_vendors: str = ""
    vt_last_analysis: str = ""
    abuse_score: str = ""
    abuse_total_reports: str = ""
    abuse_last_reported: str = ""
    otx_pulses: str = ""

    # Risk & recommendation
    overall_risk: str = "Unknown"
    recommendation: str = ""
    error: str = ""


# =============================================================================
# IOC DETECTOR
# =============================================================================

class IOCDetector:
    """Automatically detects IOC types from raw strings."""

    _IPV4 = re.compile(
        r"^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
    )
    _IPV6 = re.compile(
        r"^(([0-9a-fA-F]{1,4}:){7}[0-9a-fA-F]{1,4}|"
        r"([0-9a-fA-F]{1,4}:){1,7}:|"
        r"([0-9a-fA-F]{1,4}:){1,6}:[0-9a-fA-F]{1,4}|"
        r"([0-9a-fA-F]{1,4}:){1,5}(:[0-9a-fA-F]{1,4}){1,2}|"
        r"([0-9a-fA-F]{1,4}:){1,4}(:[0-9a-fA-F]{1,4}){1,3}|"
        r"([0-9a-fA-F]{1,4}:){1,3}(:[0-9a-fA-F]{1,4}){1,4}|"
        r"([0-9a-fA-F]{1,4}:){1,2}(:[0-9a-fA-F]{1,4}){1,5}|"
        r"[0-9a-fA-F]{1,4}:((:[0-9a-fA-F]{1,4}){1,6})|:"
        r"((:[0-9a-fA-F]{1,4}){1,7}|:))$"
    )
    _MD5 = re.compile(r"^[a-fA-F0-9]{32}$")
    _SHA1 = re.compile(r"^[a-fA-F0-9]{40}$")
    _SHA256 = re.compile(r"^[a-fA-F0-9]{64}$")

    @classmethod
    def detect(cls, ioc: str) -> str:
        """Return one of: IPv4, IPv6, Domain, URL, MD5, SHA1, SHA256, Unknown."""
        ioc = ioc.strip()
        if not ioc:
            return "Unknown"
        if cls._IPV4.match(ioc):
            return "IPv4"
        if cls._IPV6.match(ioc):
            return "IPv6"
        if cls._MD5.match(ioc):
            return "MD5"
        if cls._SHA1.match(ioc):
            return "SHA1"
        if cls._SHA256.match(ioc):
            return "SHA256"
        if ioc.lower().startswith(("http://", "https://")):
            return "URL"
        # Domain heuristic: contains at least one dot, no spaces, no path indicators
        if "." in ioc and " " not in ioc and "/" not in ioc and not ioc.startswith("."):
            # Extra validation: must look like a domain (not an IP that slipped through)
            if not cls._IPV4.match(ioc):
                return "Domain"
        return "Unknown"


# =============================================================================
# BASE API CLIENT
# =============================================================================

class BaseAPIClient:
    """Base class providing retry logic, session management, and rate-limit handling."""

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": "ThreatFusion/1.0 IOC Enrichment Tool"})

    def _request(
        self,
        method: str,
        url: str,
        params: Optional[Dict] = None,
        data: Optional[Dict] = None,
        headers: Optional[Dict] = None,
        retries: int = RETRY_ATTEMPTS
    ) -> Optional[Dict]:
        for attempt in range(1, retries + 1):
            try:
                if method.upper() == "GET":
                    resp = self.session.get(url, params=params, headers=headers, timeout=REQUEST_TIMEOUT)
                else:
                    resp = self.session.post(url, data=data, headers=headers, timeout=REQUEST_TIMEOUT)

                if resp.status_code == 429:
                    wait = 2 ** attempt
                    logger.warning(f"Rate limited ({url}). Backing off {wait}s...")
                    time.sleep(wait)
                    continue
                if resp.status_code == 200:
                    try:
                        return resp.json()
                    except ValueError:
                        return {"raw": resp.text}
                if resp.status_code in (401, 403):
                    logger.error(f"Authentication failed for {url}: HTTP {resp.status_code}")
                    return None
                logger.warning(f"HTTP {resp.status_code} from {url}")
                return None
            except requests.exceptions.Timeout:
                logger.warning(f"Timeout on {url} (attempt {attempt}/{retries})")
            except requests.exceptions.SSLError:
                logger.warning(f"SSL error on {url} (attempt {attempt}/{retries})")
            except requests.exceptions.ConnectionError:
                logger.warning(f"Connection error on {url} (attempt {attempt}/{retries})")
            except Exception as exc:
                logger.error(f"Unexpected error on {url}: {exc}")
                return None
        return None

    def _get(self, url: str, **kwargs) -> Optional[Dict]:
        return self._request("GET", url, **kwargs)

    def _post(self, url: str, **kwargs) -> Optional[Dict]:
        return self._request("POST", url, **kwargs)


# =============================================================================
# VIRUSTOTAL CLIENT
# =============================================================================

class VirusTotalClient(BaseAPIClient):
    """VirusTotal API v3 client with built-in rate limiting (4 lookups/min free tier)."""

    def __init__(self, api_key: str):
        super().__init__()
        self.api_key = api_key
        self.base_url = "https://www.virustotal.com/api/v3"
        self.session.headers.update({"x-apikey": self.api_key})
        self._last_call = 0.0
        self._min_interval = 15.0  # seconds between calls

    def _rate_limit(self):
        elapsed = time.time() - self._last_call
        if elapsed < self._min_interval:
            time.sleep(self._min_interval - elapsed)
        self._last_call = time.time()

    def _extract_vt_stats(self, data: Optional[Dict]) -> Tuple[str, str, str]:
        if not data:
            return "", "", ""
        attrs = data.get("data", {}).get("attributes", {})
        stats = attrs.get("last_analysis_stats", {})
        malicious = stats.get("malicious", 0)
        suspicious = stats.get("suspicious", 0)
        total = sum(stats.values()) if stats else 0
        score = f"{malicious}/{total}" if total else ""
        vendors = str(total)
        last_analysis = attrs.get("last_analysis_date", "")
        return score, vendors, last_analysis

    def get_ip_report(self, ip: str) -> Dict[str, str]:
        if not self.api_key:
            return {}
        self._rate_limit()
        data = self._get(f"{self.base_url}/ip-addresses/{ip}")
        score, vendors, last_analysis = self._extract_vt_stats(data)
        return {"vt_score": score, "vt_vendors": vendors, "vt_last_analysis": last_analysis}

    def get_domain_report(self, domain: str) -> Dict[str, str]:
        if not self.api_key:
            return {}
        self._rate_limit()
        data = self._get(f"{self.base_url}/domains/{domain}")
        score, vendors, last_analysis = self._extract_vt_stats(data)
        categories = ""
        if data:
            cats = data.get("data", {}).get("attributes", {}).get("categories", {})
            categories = ", ".join(f"{k}:{v}" for k, v in cats.items())
        return {
            "vt_score": score,
            "vt_vendors": vendors,
            "vt_last_analysis": last_analysis,
            "categories": categories
        }

    def get_url_report(self, url: str) -> Dict[str, str]:
        if not self.api_key:
            return {}
        self._rate_limit()
        url_id = hashlib.sha256(url.encode()).hexdigest()
        data = self._get(f"{self.base_url}/urls/{url_id}")
        score, vendors, last_analysis = self._extract_vt_stats(data)
        final_url = url
        status = ""
        if data:
            attrs = data.get("data", {}).get("attributes", {})
            final_url = attrs.get("last_final_url", url)
            status = str(attrs.get("last_http_response_code", ""))
        return {
            "vt_score": score,
            "vt_vendors": vendors,
            "last_scan": last_analysis,
            "final_url": final_url,
            "status": status
        }

    def get_file_report(self, file_hash: str) -> Dict[str, str]:
        if not self.api_key:
            return {}
        self._rate_limit()
        data = self._get(f"{self.base_url}/files/{file_hash}")
        score, vendors, last_analysis = self._extract_vt_stats(data)
        file_type = magic = names = malware_family = tags = ""
        if data:
            attrs = data.get("data", {}).get("attributes", {})
            file_type = attrs.get("type_description", "")
            magic = attrs.get("magic", "")
            names = ", ".join(attrs.get("names", [])[:5])
            popular = attrs.get("popular_threat_name", {})
            if isinstance(popular, dict):
                malware_family = ", ".join(list(popular.keys())[:3])
            tags = ", ".join(attrs.get("tags", [])[:5])
        return {
            "vt_score": score,
            "vt_vendors": vendors,
            "last_analysis": last_analysis,
            "file_type": file_type,
            "magic": magic,
            "names": names,
            "malware_family": malware_family,
            "tags": tags
        }


# =============================================================================
# ABUSEIPDB CLIENT
# =============================================================================

class AbuseIPDBClient(BaseAPIClient):
    """AbuseIPDB client with rate limiting (5 requests/min free tier)."""

    def __init__(self, api_key: str):
        super().__init__()
        self.api_key = api_key
        self.base_url = "https://api.abuseipdb.com/api/v2"
        self.session.headers.update({"Key": self.api_key, "Accept": "application/json"})
        self._last_call = 0.0
        self._min_interval = 12.0

    def _rate_limit(self):
        elapsed = time.time() - self._last_call
        if elapsed < self._min_interval:
            time.sleep(self._min_interval - elapsed)
        self._last_call = time.time()

    def check_ip(self, ip: str) -> Dict[str, str]:
        if not self.api_key:
            return {}
        self._rate_limit()
        data = self._get(f"{self.base_url}/check", params={"ipAddress": ip, "maxAgeInDays": 90})
        if not data:
            return {}
        attrs = data.get("data", {})
        return {
            "abuse_score": str(attrs.get("abuseConfidenceScore", "")),
            "abuse_total_reports": str(attrs.get("totalReports", "")),
            "abuse_last_reported": attrs.get("lastReportedAt", ""),
            "country": attrs.get("countryCode", ""),
            "isp": attrs.get("isp", ""),
            "network": attrs.get("network", "")
        }


# =============================================================================
# ALIENVAULT OTX CLIENT
# =============================================================================

class OTXClient(BaseAPIClient):
    """AlienVault OTX client."""

    def __init__(self, api_key: str):
        super().__init__()
        self.api_key = api_key
        self.base_url = "https://otx.alienvault.com/api/v1"
        if self.api_key:
            self.session.headers.update({"X-OTX-API-KEY": self.api_key})

    def get_indicator(self, indicator_type: str, indicator: str) -> Dict[str, str]:
        if not self.api_key:
            return {}
        data = self._get(f"{self.base_url}/indicators/{indicator_type}/{indicator}/general")
        if not data:
            return {}
        return {"otx_pulses": str(data.get("pulse_info", {}).get("count", 0))}


# =============================================================================
# GEOIP CLIENT
# =============================================================================

class GeoIPClient(BaseAPIClient):
    """GeoIP lookup using ip-api.com (free, no key) with optional ipgeolocation fallback."""

    def __init__(self, api_key: str = ""):
        super().__init__()
        self.api_key = api_key
        self.primary_url = "http://ip-api.com/json/{ip}"
        self.backup_url = "https://api.ipgeolocation.io/ipgeo"

    def lookup(self, ip: str) -> Dict[str, str]:
        # Primary: ip-api.com (free, no key, 45 req/min limit)
        try:
            resp = self.session.get(self.primary_url.format(ip=ip), timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                if data.get("status") == "success":
                    return {
                        "country": data.get("country", ""),
                        "city": data.get("city", ""),
                        "asn": str(data.get("as", "")),
                        "isp": data.get("isp", ""),
                        "latitude": str(data.get("lat", "")),
                        "longitude": str(data.get("lon", ""))
                    }
        except Exception as exc:
            logger.debug(f"ip-api lookup failed for {ip}: {exc}")

        # Fallback: ipgeolocation.io
        if self.api_key:
            try:
                data = self._get(self.backup_url, params={"apiKey": self.api_key, "ip": ip})
                if data:
                    return {
                        "country": data.get("country_name", ""),
                        "city": data.get("city", ""),
                        "asn": str(data.get("asn", "")),
                        "isp": data.get("isp", ""),
                        "latitude": str(data.get("latitude", "")),
                        "longitude": str(data.get("longitude", ""))
                    }
            except Exception as exc:
                logger.debug(f"ipgeolocation lookup failed for {ip}: {exc}")
        return {}


# =============================================================================
# SHODAN CLIENT
# =============================================================================

class ShodanClient:
    """Shodan wrapper with graceful degradation when key is missing."""

    def __init__(self, api_key: str):
        self.api_key = api_key
        self.api = None
        if api_key:
            try:
                self.api = shodan.Shodan(api_key)
            except Exception as exc:
                logger.error(f"Shodan initialization failed: {exc}")

    def lookup_ip(self, ip: str) -> Dict[str, str]:
        if not self.api:
            return {}
        try:
            host = self.api.host(ip)
            asn = host.get("asn", "")
            return {
                "asn": f"AS{asn}" if asn else "",
                "isp": host.get("isp", ""),
                "network": host.get("network", "")
            }
        except Exception as exc:
            logger.debug(f"Shodan lookup failed for {ip}: {exc}")
            return {}


# =============================================================================
# WHOIS CLIENT
# =============================================================================

class WHOISClient:
    """WHOIS lookup wrapper with robust date normalization."""

    @staticmethod
    def _normalize_date(value) -> str:
        if value is None:
            return ""
        if isinstance(value, list):
            value = value[0]
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        return str(value)

    def lookup(self, domain: str) -> Dict[str, str]:
        try:
            w = whois.whois(domain)
            return {
                "registrar": str(w.registrar) if w.registrar else "",
                "creation_date": self._normalize_date(w.creation_date),
                "expiration_date": self._normalize_date(w.expiration_date)
            }
        except Exception as exc:
            logger.debug(f"WHOIS failed for {domain}: {exc}")
            return {}


# =============================================================================
# RISK ENGINE
# =============================================================================

class RiskEngine:
    """
    Multi-source risk scoring engine.
    Weights: VirusTotal (50%), AbuseIPDB (30%), OTX (20%).
    """

    @staticmethod
    def _parse_vt(score: str) -> int:
        if not score or "/" not in score:
            return 0
        try:
            return int(score.split("/")[0])
        except ValueError:
            return 0

    @staticmethod
    def _parse_abuse(score: str) -> int:
        if not score:
            return 0
        try:
            return int(score)
        except ValueError:
            return 0

    @classmethod
    def calculate(cls, ioc_type: str, vt_score: str, abuse_score: str, otx_pulses: str) -> Tuple[str, str]:
        vt_malicious = cls._parse_vt(vt_score)
        abuse = cls._parse_abuse(abuse_score)
        otx = int(otx_pulses) if otx_pulses and otx_pulses.isdigit() else 0

        # VT level (0-5 scale mapped to 0,2,3,4,5)
        if vt_malicious == 0:
            vt_level = 0
        elif 1 <= vt_malicious <= 4:
            vt_level = 2
        elif 5 <= vt_malicious <= 9:
            vt_level = 3
        elif 10 <= vt_malicious <= 19:
            vt_level = 4
        else:
            vt_level = 5

        # Abuse level (0-5 scale)
        if abuse == 0:
            abuse_level = 0
        elif 1 <= abuse <= 20:
            abuse_level = 1
        elif 21 <= abuse <= 50:
            abuse_level = 2
        elif 51 <= abuse <= 80:
            abuse_level = 4
        else:
            abuse_level = 5

        # OTX level
        if otx == 0:
            otx_level = 0
        elif otx <= 2:
            otx_level = 1
        elif otx <= 5:
            otx_level = 2
        else:
            otx_level = 3

        # Combined weighted score (0-100 scale approximation)
        combined = (vt_level * 10) + (abuse_level * 6) + (otx_level * 4)

        if combined == 0:
            overall = "Clean"
            recommendation = "No malicious indicators detected. No action required; maintain standard monitoring."
        elif combined <= 20:
            overall = "Low"
            recommendation = "Low risk profile. Verify business context before action. Consider adding to watchlist."
        elif combined <= 40:
            overall = "Suspicious"
            recommendation = "Suspicious indicators present. Correlate with internal logs and investigate origin."
        elif combined <= 60:
            overall = "Medium"
            recommendation = "Medium risk. Block at perimeter if not business-critical. Hunt for related IOCs."
        elif combined <= 80:
            overall = "High"
            recommendation = "High risk. Immediate containment advised. Block IOC and review affected assets."
        else:
            overall = "Critical"
            recommendation = "Critical threat confirmed. Block immediately, isolate systems, and initiate IR procedures."

        return overall, recommendation


# =============================================================================
# ENRICHER ORCHESTRATOR
# =============================================================================

class IOCEnricher:
    """Orchestrates all API clients to produce a fully enriched IOC record."""

    def __init__(self):
        self.vt = VirusTotalClient(VT_API_KEY)
        self.abuse = AbuseIPDBClient(ABUSE_API_KEY)
        self.otx = OTXClient(OTX_API_KEY)
        self.geo = GeoIPClient(GEOIP_API_KEY)
        self.shodan = ShodanClient(SHODAN_API_KEY)
        self.whois = WHOISClient()
        self.risk = RiskEngine()

    def enrich(self, ioc: str) -> EnrichedIOC:
        ioc_type = IOCDetector.detect(ioc)
        result = EnrichedIOC(ioc=ioc, ioc_type=ioc_type)

        if ioc_type == "Unknown":
            result.error = "Unable to automatically detect IOC type"
            result.overall_risk = "Unknown"
            result.recommendation = "Manual review required."
            return result

        try:
            if ioc_type in ("IPv4", "IPv6"):
                result = self._enrich_ip(ioc, result)
            elif ioc_type == "Domain":
                result = self._enrich_domain(ioc, result)
            elif ioc_type == "URL":
                result = self._enrich_url(ioc, result)
            elif ioc_type in ("MD5", "SHA1", "SHA256"):
                result = self._enrich_hash(ioc, result)

            result.overall_risk, result.recommendation = self.risk.calculate(
                ioc_type, result.vt_score, result.abuse_score, result.otx_pulses
            )
        except Exception as exc:
            logger.error(f"Enrichment pipeline failed for {ioc}: {exc}")
            result.error = str(exc)
            result.overall_risk = "Error"
            result.recommendation = "Manual review required due to processing error."

        return result

    def _enrich_ip(self, ip: str, result: EnrichedIOC) -> EnrichedIOC:
        # VirusTotal
        vt = self.vt.get_ip_report(ip)
        result.vt_score = vt.get("vt_score", "")
        result.vt_vendors = vt.get("vt_vendors", "")
        result.vt_last_analysis = vt.get("vt_last_analysis", "")

        # AbuseIPDB
        abuse = self.abuse.check_ip(ip)
        result.abuse_score = abuse.get("abuse_score", "")
        result.abuse_total_reports = abuse.get("abuse_total_reports", "")
        result.abuse_last_reported = abuse.get("abuse_last_reported", "")
        result.country = abuse.get("country", "")
        result.isp = abuse.get("isp", "")
        result.network = abuse.get("network", "")

        # GeoIP
        geo = self.geo.lookup(ip)
        if not result.country:
            result.country = geo.get("country", "")
        result.city = geo.get("city", "")
        result.asn = geo.get("asn", "")
        if not result.isp:
            result.isp = geo.get("isp", "")
        result.latitude = geo.get("latitude", "")
        result.longitude = geo.get("longitude", "")

        # Shodan
        shodan_data = self.shodan.lookup_ip(ip)
        if not result.asn:
            result.asn = shodan_data.get("asn", "")
        if not result.isp:
            result.isp = shodan_data.get("isp", "")
        if not result.network:
            result.network = shodan_data.get("network", "")

        # OTX
        otx = self.otx.get_indicator("IPv4", ip)
        result.otx_pulses = otx.get("otx_pulses", "")

        return result

    def _enrich_domain(self, domain: str, result: EnrichedIOC) -> EnrichedIOC:
        # VirusTotal
        vt = self.vt.get_domain_report(domain)
        result.vt_score = vt.get("vt_score", "")
        result.vt_vendors = vt.get("vt_vendors", "")
        result.categories = vt.get("categories", "")

        # WHOIS
        w = self.whois.lookup(domain)
        result.registrar = w.get("registrar", "")
        result.creation_date = w.get("creation_date", "")
        result.expiration_date = w.get("expiration_date", "")

        # OTX
        otx = self.otx.get_indicator("domain", domain)
        result.otx_pulses = otx.get("otx_pulses", "")

        # Attempt GeoIP via DNS resolution
        try:
            resolved = socket.gethostbyname(domain)
            geo = self.geo.lookup(resolved)
            result.country = geo.get("country", "")
        except Exception:
            pass

        return result

    def _enrich_url(self, url: str, result: EnrichedIOC) -> EnrichedIOC:
        # VirusTotal
        vt = self.vt.get_url_report(url)
        result.vt_score = vt.get("vt_score", "")
        result.vt_vendors = vt.get("vt_vendors", "")
        result.last_scan = vt.get("last_scan", "")
        result.final_url = vt.get("final_url", url)
        result.status = vt.get("status", "")

        # OTX
        otx = self.otx.get_indicator("url", url)
        result.otx_pulses = otx.get("otx_pulses", "")

        # Category heuristic
        result.category = "Uncategorized"

        return result

    def _enrich_hash(self, file_hash: str, result: EnrichedIOC) -> EnrichedIOC:
        # VirusTotal
        vt = self.vt.get_file_report(file_hash)
        result.vt_score = vt.get("vt_score", "")
        result.vt_vendors = vt.get("vt_vendors", "")
        result.file_type = vt.get("file_type", "")
        result.magic = vt.get("magic", "")
        result.names = vt.get("names", "")
        result.malware_family = vt.get("malware_family", "")
        result.tags = vt.get("tags", "")
        result.vt_last_analysis = vt.get("last_analysis", "")

        # OTX
        otx = self.otx.get_indicator("file", file_hash)
        result.otx_pulses = otx.get("otx_pulses", "")

        return result


# =============================================================================
# OUTPUT GENERATORS
# =============================================================================

class OutputGenerator:
    """Handles CSV, Excel, HTML, and PDF report generation."""

    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def to_csv(self, results: List[EnrichedIOC], filename: str = "IOC_Report.csv"):
        filepath = os.path.join(self.output_dir, filename)
        if not results:
            logger.warning("No results to write to CSV.")
            return
        df = pd.DataFrame([asdict(r) for r in results])
        df.to_csv(filepath, index=False, encoding="utf-8")
        logger.info(f"CSV report saved: {filepath}")

    def to_excel(self, results: List[EnrichedIOC], filename: str = "IOC_Report.xlsx"):
        filepath = os.path.join(self.output_dir, filename)
        if not results:
            logger.warning("No results to write to Excel.")
            return

        df = pd.DataFrame([asdict(r) for r in results])
        wb = Workbook()
        ws = wb.active
        ws.title = "IOC Report"

        headers = list(df.columns)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Data rows
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                # Risk-based conditional coloring
                if headers[col_idx - 1] == "overall_risk":
                    if value == "Critical":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif value == "High":
                        cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif value == "Medium":
                        cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
                    elif value == "Suspicious":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "Low":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "Clean":
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # Auto-width with cap
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Enable filters
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        wb.save(filepath)
        logger.info(f"Excel report saved: {filepath}")

    def to_html(self, results: List[EnrichedIOC], filename: str = "IOC_Report.html"):
        filepath = os.path.join(self.output_dir, filename)
        if not results:
            logger.warning("No results to write to HTML.")
            return

        total = len(results)
        malicious = sum(1 for r in results if r.overall_risk in ("Critical", "High"))
        suspicious = sum(1 for r in results if r.overall_risk == "Suspicious")
        clean = sum(1 for r in results if r.overall_risk in ("Clean", "Low"))

        type_dist: Dict[str, int] = {}
        country_dist: Dict[str, int] = {}
        risk_dist: Dict[str, int] = {}
        asn_counts: Dict[str, int] = {}
        isp_counts: Dict[str, int] = {}

        for r in results:
            type_dist[r.ioc_type] = type_dist.get(r.ioc_type, 0) + 1
            if r.country:
                country_dist[r.country] = country_dist.get(r.country, 0) + 1
            risk_dist[r.overall_risk] = risk_dist.get(r.overall_risk, 0) + 1
            if r.asn:
                asn_counts[r.asn] = asn_counts.get(r.asn, 0) + 1
            if r.isp:
                isp_counts[r.isp] = isp_counts.get(r.isp, 0) + 1

        top_asn = sorted(asn_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        top_isp = sorted(isp_counts.items(), key=lambda x: x[1], reverse=True)[:5]

        type_labels = list(type_dist.keys())
        type_values = list(type_dist.values())
        risk_labels = list(risk_dist.keys())
        risk_values = list(risk_dist.values())
        country_labels = list(country_dist.keys())[:10]
        country_values = list(country_dist.values())[:10]
        asn_labels = [x[0] for x in top_asn]
        asn_values = [x[1] for x in top_asn]
        isp_labels = [x[0] for x in top_isp]
        isp_values = [x[1] for x in top_isp]

        template = Template("""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ThreatFusion - IOC Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
:root {
  --bg-dark: #0b1120;
  --bg-card: #1e293b;
  --text-primary: #f8fafc;
  --text-secondary: #94a3b8;
  --accent-blue: #3b82f6;
  --accent-green: #10b981;
  --accent-yellow: #f59e0b;
  --accent-red: #ef4444;
  --accent-orange: #f97316;
  --border: #334155;
}
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
  font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
  background: var(--bg-dark);
  color: var(--text-primary);
  line-height: 1.6;
}
.container { max-width: 1440px; margin: 0 auto; padding: 24px; }
header {
  text-align: center;
  padding: 32px 0;
  border-bottom: 2px solid var(--accent-blue);
  margin-bottom: 32px;
}
header h1 { font-size: 2.8rem; color: var(--accent-blue); text-transform: uppercase; letter-spacing: 3px; font-weight: 800; }
header p { color: var(--text-secondary); margin-top: 8px; font-size: 1rem; }
.summary-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
  gap: 20px;
  margin-bottom: 32px;
}
.card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 24px;
  text-align: center;
  transition: transform 0.2s, border-color 0.2s;
}
.card:hover { transform: translateY(-4px); border-color: var(--accent-blue); }
.card h3 { font-size: 2.2rem; margin-bottom: 6px; font-weight: 700; }
.card p { color: var(--text-secondary); font-size: 0.85rem; text-transform: uppercase; letter-spacing: 1px; }
.card.total h3 { color: var(--accent-blue); }
.card.malicious h3 { color: var(--accent-red); }
.card.suspicious h3 { color: var(--accent-orange); }
.card.clean h3 { color: var(--accent-green); }
.charts-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(380px, 1fr));
  gap: 24px;
  margin-bottom: 32px;
}
.chart-container {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 20px;
}
.chart-container h3 { margin-bottom: 16px; color: var(--text-primary); font-size: 1.05rem; font-weight: 600; }
.search-box {
  width: 100%;
  padding: 14px 20px;
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: 10px;
  color: var(--text-primary);
  font-size: 1rem;
  margin-bottom: 24px;
  transition: border-color 0.2s;
}
.search-box:focus { outline: none; border-color: var(--accent-blue); }
table {
  width: 100%;
  border-collapse: collapse;
  background: var(--bg-card);
  border-radius: 12px;
  overflow: hidden;
  font-size: 0.9rem;
}
th, td {
  padding: 12px 14px;
  text-align: left;
  border-bottom: 1px solid var(--border);
}
th {
  background: #0f172a;
  color: var(--accent-blue);
  text-transform: uppercase;
  font-size: 0.75rem;
  letter-spacing: 1px;
  cursor: pointer;
  position: sticky;
  top: 0;
  user-select: none;
}
th:hover { background: #1e293b; }
tr:hover { background: rgba(59, 130, 246, 0.08); }
.risk-critical { color: var(--accent-red); font-weight: 700; }
.risk-high { color: var(--accent-orange); font-weight: 700; }
.risk-medium { color: var(--accent-yellow); font-weight: 700; }
.risk-suspicious { color: #fb923c; font-weight: 700; }
.risk-low { color: var(--accent-green); }
.risk-clean { color: #34d399; }
.risk-unknown { color: var(--text-secondary); }
.timestamp { text-align: center; color: var(--text-secondary); margin-top: 32px; font-size: 0.85rem; }
@media (max-width: 768px) {
  .charts-grid { grid-template-columns: 1fr; }
  header h1 { font-size: 1.8rem; }
}
</style>
</head>
<body>
<div class="container">
  <header>
    <h1>ThreatFusion</h1>
    <p>IOC Enrichment Dashboard &mdash; Generated on {{ timestamp }}</p>
  </header>

  <div class="summary-grid">
    <div class="card total"><h3>{{ total }}</h3><p>Total IOCs</p></div>
    <div class="card malicious"><h3>{{ malicious }}</h3><p>Malicious</p></div>
    <div class="card suspicious"><h3>{{ suspicious }}</h3><p>Suspicious</p></div>
    <div class="card clean"><h3>{{ clean }}</h3><p>Clean / Low</p></div>
  </div>

  <div class="charts-grid">
    <div class="chart-container"><h3>IOC Type Distribution</h3><canvas id="typeChart"></canvas></div>
    <div class="chart-container"><h3>Risk Distribution</h3><canvas id="riskChart"></canvas></div>
    <div class="chart-container"><h3>Country Distribution (Top 10)</h3><canvas id="countryChart"></canvas></div>
    <div class="chart-container"><h3>Top ASN</h3><canvas id="asnChart"></canvas></div>
    <div class="chart-container"><h3>Top ISP</h3><canvas id="ispChart"></canvas></div>
  </div>

  <input type="text" class="search-box" id="tableSearch" placeholder="Search IOCs..." onkeyup="filterTable()">

  <div style="overflow-x: auto;">
    <table id="iocTable">
      <thead>
        <tr>
          <th onclick="sortTable(0)">IOC</th>
          <th onclick="sortTable(1)">Type</th>
          <th onclick="sortTable(2)">Risk</th>
          <th onclick="sortTable(3)">VT Score</th>
          <th onclick="sortTable(4)">Abuse Score</th>
          <th onclick="sortTable(5)">OTX Pulses</th>
          <th onclick="sortTable(6)">Country</th>
          <th onclick="sortTable(7)">ISP / Registrar</th>
          <th onclick="sortTable(8)">Recommendation</th>
        </tr>
      </thead>
      <tbody>
        {% for r in results %}
        <tr>
          <td>{{ r.ioc }}</td>
          <td>{{ r.ioc_type }}</td>
          <td class="risk-{{ r.overall_risk|lower }}">{{ r.overall_risk }}</td>
          <td>{{ r.vt_score }}</td>
          <td>{{ r.abuse_score }}</td>
          <td>{{ r.otx_pulses }}</td>
          <td>{{ r.country }}</td>
          <td>{{ r.isp if r.isp else r.registrar }}</td>
          <td>{{ r.recommendation }}</td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>

  <p class="timestamp">Report generated by ThreatFusion v1.0</p>
</div>

<script>
const chartColors = ['#3b82f6','#10b981','#f59e0b','#ef4444','#f97316','#8b5cf6','#06b6d4','#ec4899'];

new Chart(document.getElementById('typeChart'), {
  type: 'doughnut',
  data: { labels: {{ type_labels|tojson }}, datasets: [{ data: {{ type_values|tojson }}, backgroundColor: chartColors, borderWidth: 0 }] },
  options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'right', labels: { color: '#f1f5f9', font: { size: 12 } } } } }
});

new Chart(document.getElementById('riskChart'), {
  type: 'bar',
  data: { labels: {{ risk_labels|tojson }}, datasets: [{ label: 'Count', data: {{ risk_values|tojson }}, backgroundColor: chartColors, borderRadius: 6 }] },
  options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { color: '#94a3b8' }, grid: { color: '#334155' } }, x: { ticks: { color: '#94a3b8' }, grid: { color: '#334155' } } } }
});

new Chart(document.getElementById('countryChart'), {
  type: 'bar',
  data: { labels: {{ country_labels|tojson }}, datasets: [{ label: 'Count', data: {{ country_values|tojson }}, backgroundColor: '#3b82f6', borderRadius: 6 }] },
  options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { color: '#94a3b8' }, grid: { color: '#334155' } }, x: { ticks: { color: '#94a3b8', autoSkip: false, maxRotation: 45, minRotation: 45 }, grid: { color: '#334155' } } } }
});

new Chart(document.getElementById('asnChart'), {
  type: 'bar',
  data: { labels: {{ asn_labels|tojson }}, datasets: [{ label: 'Count', data: {{ asn_values|tojson }}, backgroundColor: '#f97316', borderRadius: 6 }] },
  options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { color: '#94a3b8' }, grid: { color: '#334155' } }, x: { ticks: { color: '#94a3b8', autoSkip: false, maxRotation: 45, minRotation: 45 }, grid: { color: '#334155' } } } }
});

new Chart(document.getElementById('ispChart'), {
  type: 'bar',
  data: { labels: {{ isp_labels|tojson }}, datasets: [{ label: 'Count', data: {{ isp_values|tojson }}, backgroundColor: '#8b5cf6', borderRadius: 6 }] },
  options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { ticks: { color: '#94a3b8' }, grid: { color: '#334155' } }, x: { ticks: { color: '#94a3b8', autoSkip: false, maxRotation: 45, minRotation: 45 }, grid: { color: '#334155' } } } }
});

function filterTable() {
  const input = document.getElementById('tableSearch');
  const filter = input.value.toUpperCase();
  const table = document.getElementById('iocTable');
  const tr = table.getElementsByTagName('tr');
  for (let i = 1; i < tr.length; i++) {
    const td = tr[i].getElementsByTagName('td');
    let visible = false;
    for (let j = 0; j < td.length; j++) {
      if (td[j] && td[j].textContent.toUpperCase().indexOf(filter) > -1) { visible = true; break; }
    }
    tr[i].style.display = visible ? '' : 'none';
  }
}

function sortTable(n) {
  const table = document.getElementById('iocTable');
  let rows = Array.from(table.rows).slice(1);
  const dir = table.getAttribute('data-dir') === 'asc' ? 'desc' : 'asc';
  table.setAttribute('data-dir', dir);
  rows.sort((a, b) => {
    const x = a.cells[n].textContent.trim();
    const y = b.cells[n].textContent.trim();
    return dir === 'asc' ? x.localeCompare(y) : y.localeCompare(x);
  });
  rows.forEach(row => table.tBodies[0].appendChild(row));
}
</script>
</body>
</html>
        """)

        html_content = template.render(
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            total=total,
            malicious=malicious,
            suspicious=suspicious,
            clean=clean,
            results=results,
            type_labels=type_labels,
            type_values=type_values,
            risk_labels=risk_labels,
            risk_values=risk_values,
            country_labels=country_labels,
            country_values=country_values,
            asn_labels=asn_labels,
            asn_values=asn_values,
            isp_labels=isp_labels,
            isp_values=isp_values
        )

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)
        logger.info(f"HTML dashboard saved: {filepath}")

    def to_pdf(self, results: List[EnrichedIOC], filename: str = "IOC_Report.pdf"):
        filepath = os.path.join(self.output_dir, filename)
        if not results:
            logger.warning("No results to write to PDF.")
            return

        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=30
        )
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1F4E78"),
            spaceAfter=18,
            alignment=1,
            fontName="Helvetica-Bold"
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#1F4E78"),
            spaceAfter=10,
            spaceBefore=16,
            fontName="Helvetica-Bold"
        )
        body_style = ParagraphStyle(
            "CustomBody",
            parent=styles["BodyText"],
            fontSize=9,
            leading=13,
            alignment=0
        )
        bullet_style = ParagraphStyle(
            "Bullet",
            parent=body_style,
            leftIndent=20,
            bulletIndent=10,
            spaceAfter=4
        )

        # Title
        story.append(Paragraph("ThreatFusion - IOC Enrichment Report", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", body_style))
        story.append(Spacer(1, 20))

        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        total = len(results)
        malicious = sum(1 for r in results if r.overall_risk in ("Critical", "High"))
        suspicious = sum(1 for r in results if r.overall_risk == "Suspicious")
        clean = sum(1 for r in results if r.overall_risk in ("Clean", "Low"))
        unknown = sum(1 for r in results if r.overall_risk in ("Unknown", "Error"))

        summary_text = f"""
        <b>Total IOCs Processed:</b> {total}<br/>
        <b>Malicious (Critical / High):</b> {malicious}<br/>
        <b>Suspicious:</b> {suspicious}<br/>
        <b>Clean / Low Risk:</b> {clean}<br/>
        <b>Unknown / Error:</b> {unknown}<br/><br/>
        This report provides a comprehensive analysis of indicators of compromise (IOCs) enriched
        via VirusTotal, AbuseIPDB, AlienVault OTX, WHOIS, GeoIP, and Shodan. Immediate attention
        is required for all IOCs classified as <b>Critical</b> or <b>High</b> risk.
        """
        story.append(Paragraph(summary_text, body_style))
        story.append(Spacer(1, 14))

        # IOC Summary by Type
        story.append(Paragraph("IOC Summary by Type", heading_style))
        type_counts: Dict[str, int] = {}
        for r in results:
            type_counts[r.ioc_type] = type_counts.get(r.ioc_type, 0) + 1
        type_data = [["IOC Type", "Count"]] + [[k, v] for k, v in sorted(type_counts.items())]
        type_table = Table(type_data, colWidths=[300, 100])
        type_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
        ]))
        story.append(type_table)
        story.append(Spacer(1, 14))

        # Risk Distribution
        story.append(Paragraph("Risk Distribution", heading_style))
        risk_data = [["Risk Level", "Count"]] + [[k, v] for k, v in sorted(risk_dist.items(), key=lambda x: x[1], reverse=True)]
        risk_table = Table(risk_data, colWidths=[300, 100])
        risk_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
        ]))
        story.append(risk_table)
        story.append(Spacer(1, 14))

        # Recommendations
        story.append(Paragraph("Strategic Recommendations", heading_style))
        recs = []
        if malicious > 0:
            recs.append("Immediately block all Critical and High-risk IOCs at perimeter firewalls, proxies, and endpoint protection platforms.")
        if suspicious > 0:
            recs.append("Investigate Suspicious IOCs in sandboxed environments before allowing network access.")
        if clean == total:
            recs.append("No immediate threats detected. Maintain standard monitoring and detection procedures.")
        recs.append("Integrate ThreatFusion enrichment feeds into SIEM for continuous monitoring and automated alerting.")
        recs.append("Review and update threat intelligence sources quarterly to ensure coverage of emerging threat actors.")
        recs.append("For hash-based IOCs, deploy YARA rules or host-based signatures to detect related malware variants.")

        for rec in recs:
            story.append(Paragraph(f"&bull; {rec}", bullet_style))
        story.append(PageBreak())

        # Detailed IOC Table
        story.append(Paragraph("Detailed IOC Analysis", heading_style))
        table_data = [[
            "IOC", "Type", "Risk", "VT Score", "Abuse", "OTX", "Country", "Recommendation"
        ]]
        for r in results:
            table_data.append([
                r.ioc[:45],
                r.ioc_type,
                r.overall_risk,
                r.vt_score,
                r.abuse_score,
                r.otx_pulses,
                r.country[:18] if r.country else "",
                r.recommendation[:90]
            ])

        detail_table = Table(table_data, colWidths=[95, 48, 52, 48, 42, 38, 65, 135], repeatRows=1)
        detail_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("WORDWRAP", (0, 0), (-1, -1), True),
        ]))

        # Row-level risk coloring
        for i, r in enumerate(results, 1):
            if r.overall_risk == "Critical":
                detail_table.setStyle(TableStyle([("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFCCCC"))]))
            elif r.overall_risk == "High":
                detail_table.setStyle(TableStyle([("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFE5CC"))]))
            elif r.overall_risk == "Medium":
                detail_table.setStyle(TableStyle([("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFFFCC"))]))
            elif r.overall_risk == "Suspicious":
                detail_table.setStyle(TableStyle([("BACKGROUND", (0, i), (-1, i), colors.HexColor("#E6F3FF"))]))

        story.append(detail_table)
        doc.build(story)
        logger.info(f"PDF report saved: {filepath}")


# =============================================================================
# MAIN CONTROLLER
# =============================================================================

class ThreatFusion:
    """Main application controller."""

    def __init__(self):
        self.enricher = IOCEnricher()
        self.output = OutputGenerator(OUTPUT_DIR)
        self.results: List[EnrichedIOC] = []

    def read_iocs(self, filepath: str) -> List[str]:
        """Read IOCs from file, deduplicate, and filter blanks."""
        if not os.path.exists(filepath):
            logger.error(f"Input file not found: {filepath}")
            return []
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                raw = [line.strip() for line in f if line.strip() and not line.strip().startswith("#")]
            # Preserve order, case-insensitive deduplication
            seen = set()
            unique = []
            for ioc in raw:
                key = ioc.lower()
                if key not in seen:
                    seen.add(key)
                    unique.append(ioc)
            logger.info(f"Loaded {len(unique)} unique IOCs from {filepath}")
            return unique
        except Exception as exc:
            logger.error(f"Failed to read IOC file: {exc}")
            return []

    def process_single(self, ioc: str) -> EnrichedIOC:
        logger.info(f"Processing IOC: {ioc}")
        return self.enricher.enrich(ioc)

    def run(self):
        start_time = time.time()
        logger.info("=" * 70)
        logger.info("ThreatFusion - IOC Enrichment Tool v1.0")
        logger.info("=" * 70)

        iocs = self.read_iocs(INPUT_FILE)
        if not iocs:
            logger.error("No IOCs to process. Exiting.")
            print("[ERROR] No IOCs found. Please create 'IOC List.txt' with one IOC per line.")
            return

        # Parallel enrichment with progress bar
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_ioc = {executor.submit(self.process_single, ioc): ioc for ioc in iocs}
            for future in tqdm(as_completed(future_to_ioc), total=len(iocs), desc="Enriching IOCs", unit="ioc"):
                try:
                    result = future.result()
                    self.results.append(result)
                except Exception as exc:
                    ioc = future_to_ioc[future]
                    logger.error(f"Unhandled exception for {ioc}: {exc}")
                    self.results.append(EnrichedIOC(ioc=ioc, ioc_type="Unknown", error=str(exc)))

        # Sort by severity (most critical first)
        severity_map = {
            "Critical": 0, "High": 1, "Medium": 2, "Suspicious": 3,
            "Low": 4, "Clean": 5, "Unknown": 6, "Error": 7
        }
        self.results.sort(key=lambda x: severity_map.get(x.overall_risk, 99))

        # Generate all outputs
        logger.info("Generating reports...")
        self.output.to_csv(self.results)
        self.output.to_excel(self.results)
        self.output.to_html(self.results)
        self.output.to_pdf(self.results)

        # Final summary
        elapsed = time.time() - start_time
        total = len(self.results)
        ips = sum(1 for r in self.results if r.ioc_type in ("IPv4", "IPv6"))
        domains = sum(1 for r in self.results if r.ioc_type == "Domain")
        urls = sum(1 for r in self.results if r.ioc_type == "URL")
        hashes = sum(1 for r in self.results if r.ioc_type in ("MD5", "SHA1", "SHA256"))
        malicious = sum(1 for r in self.results if r.overall_risk in ("Critical", "High"))
        suspicious = sum(1 for r in self.results if r.overall_risk == "Suspicious")
        clean = sum(1 for r in self.results if r.overall_risk in ("Clean", "Low"))

        summary_lines = [
            "=" * 70,
            "ENRICHMENT COMPLETE",
            "=" * 70,
            f"  Total IOCs:       {total}",
            f"  IPs:              {ips}",
            f"  Domains:          {domains}",
            f"  URLs:             {urls}",
            f"  Hashes:           {hashes}",
            f"  Malicious:        {malicious}",
            f"  Suspicious:       {suspicious}",
            f"  Clean / Low:      {clean}",
            f"  Execution Time:   {elapsed:.2f}s",
            f"  Reports saved to: {os.path.abspath(OUTPUT_DIR)}",
            "=" * 70,
        ]
        for line in summary_lines:
            logger.info(line)

        print()
        for line in summary_lines:
            print(line)
        print()


if __name__ == "__main__":
    app = ThreatFusion()
    app.run()
